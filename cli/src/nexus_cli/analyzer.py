"""Nexus analyzer - Core calculation logic for Python CLI.

This module implements nexus detection logic that mirrors the
TypeScript implementation in the web app.
"""

import pandas as pd
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

from .config_loader import get_loader, StateRule


@dataclass
class Transaction:
    """A single transaction record."""

    id: str
    state_code: str
    amount: float
    date: date
    revenue_type: Optional[str] = None


@dataclass
class StateResult:
    """Nexus analysis result for a single state."""

    state_code: str
    has_nexus: bool
    total_revenue: float
    total_transactions: int
    threshold_revenue: float
    threshold_transactions: Optional[int]
    breach_type: Optional[str]  # 'revenue' or 'transactions'
    breach_date: Optional[date]
    breach_transaction_id: Optional[str]
    threshold_percentage: float
    tax_rate: float
    estimated_liability: float


@dataclass
class AnalysisResult:
    """Complete analysis result for all states."""

    state_results: List[StateResult]
    total_liability: float
    analysis_year: int
    transaction_count: int
    warnings: List[str]


class NexusAnalyzer:
    """Analyzes transaction data for nexus obligations."""

    def __init__(self):
        """Initialize the analyzer with configuration."""
        self.loader = get_loader()

    def load_csv(self, file_path: Path) -> List[Transaction]:
        """Load transactions from a CSV file.

        Args:
            file_path: Path to the CSV file

        Returns:
            List of Transaction objects

        Raises:
            ValueError: If required columns are missing or data is invalid
        """
        try:
            df = pd.read_csv(file_path)
        except Exception as e:
            raise ValueError(f"Failed to read CSV file: {e}")

        # Check required columns
        required_cols = ["date", "state", "amount"]
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            # Try alternative column names
            col_mapping = {
                "date": ["transaction_date", "sale_date", "order_date"],
                "state": ["state_code", "State", "STATE"],
                "amount": ["sale_amount", "sales_amount", "revenue", "total"],
            }

            for col in missing_cols:
                found = False
                for alt_name in col_mapping.get(col, []):
                    if alt_name in df.columns:
                        df.rename(columns={alt_name: col}, inplace=True)
                        found = True
                        break

                if not found:
                    raise ValueError(f"Missing required column: {col}")

        # Convert to transactions
        transactions = []
        for idx, row in df.iterrows():
            try:
                # Parse date
                if isinstance(row["date"], str):
                    txn_date = datetime.strptime(row["date"], "%Y-%m-%d").date()
                else:
                    txn_date = pd.to_datetime(row["date"]).date()

                # Parse state code
                state_code = str(row["state"]).strip().upper()
                if len(state_code) > 2:
                    # Might be full state name, try to map it
                    state_code = self._map_state_name(state_code)

                # Parse amount
                amount = float(row["amount"])

                # Create transaction
                txn = Transaction(
                    id=str(row.get("id", f"tx-{idx}")),
                    state_code=state_code,
                    amount=amount,
                    date=txn_date,
                    revenue_type=row.get("revenue_type"),
                )

                transactions.append(txn)

            except Exception as e:
                raise ValueError(f"Error parsing row {idx}: {e}")

        return transactions

    def analyze(
        self,
        transactions: List[Transaction],
        year: Optional[int] = None,
        ignore_marketplace: bool = False,
        include_negatives: bool = False,
    ) -> AnalysisResult:
        """Analyze transactions for nexus.

        Args:
            transactions: List of transactions to analyze
            year: Year to analyze (default: current year)
            ignore_marketplace: Whether to ignore marketplace transactions
            include_negatives: Whether to include negative amounts

        Returns:
            AnalysisResult with findings for each state
        """
        if year is None:
            year = date.today().year

        warnings = []

        # Filter transactions
        filtered = self._filter_transactions(
            transactions, year, ignore_marketplace, include_negatives, warnings
        )

        # Group by state
        by_state = self._group_by_state(filtered)

        # Analyze each state
        state_results = []
        for state_code, state_txns in by_state.items():
            rule = self.loader.get_rule_for_state(state_code)
            if rule is None:
                warnings.append(f"No rule found for state: {state_code}")
                continue

            result = self._analyze_state(state_code, state_txns, rule)
            state_results.append(result)

        # Add states with no transactions
        all_rules = self.loader.get_current_rules()
        processed_states = {r.state_code for r in state_results}

        for rule in all_rules:
            if rule.state_code not in processed_states:
                tax_rate = self.loader.get_tax_rate(rule.state_code) or 0.0
                state_results.append(
                    StateResult(
                        state_code=rule.state_code,
                        has_nexus=False,
                        total_revenue=0.0,
                        total_transactions=0,
                        threshold_revenue=rule.amount,
                        threshold_transactions=rule.transactions,
                        breach_type=None,
                        breach_date=None,
                        breach_transaction_id=None,
                        threshold_percentage=0.0,
                        tax_rate=tax_rate,
                        estimated_liability=0.0,
                    )
                )

        # Calculate total liability
        total_liability = sum(r.estimated_liability for r in state_results)

        return AnalysisResult(
            state_results=state_results,
            total_liability=total_liability,
            analysis_year=year,
            transaction_count=len(filtered),
            warnings=warnings,
        )

    def export_excel(self, result: AnalysisResult, output_path: Path) -> None:
        """Export analysis results to Excel.

        Args:
            result: Analysis result to export
            output_path: Path for output Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        # Create workbook
        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Header
        ws_summary["A1"] = "Nexus Analysis Report"
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary["A2"] = f"Analysis Year: {result.analysis_year}"
        ws_summary["A3"] = f"Generated: {date.today().strftime('%Y-%m-%d')}"

        # Summary stats
        ws_summary["A5"] = "Summary"
        ws_summary["A5"].font = Font(size=14, bold=True)

        nexus_count = len([r for r in result.state_results if r.has_nexus])
        ws_summary["A6"] = "States with Nexus:"
        ws_summary["B6"] = nexus_count

        ws_summary["A7"] = "Total Liability:"
        ws_summary["B7"] = result.total_liability
        ws_summary["B7"].number_format = "$#,##0.00"

        ws_summary["A8"] = "Transactions Analyzed:"
        ws_summary["B8"] = result.transaction_count

        # Nexus states detail
        ws_detail = wb.create_sheet("Nexus States")

        headers = [
            "State",
            "Revenue",
            "Threshold",
            "% of Threshold",
            "Transactions",
            "Breach Type",
            "Breach Date",
            "Tax Rate",
            "Est. Liability",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add nexus states
        row = 2
        for state in sorted(
            [r for r in result.state_results if r.has_nexus],
            key=lambda x: x.total_revenue,
            reverse=True,
        ):
            ws_detail.cell(row=row, column=1, value=state.state_code)
            ws_detail.cell(row=row, column=2, value=state.total_revenue).number_format = "$#,##0.00"
            ws_detail.cell(row=row, column=3, value=state.threshold_revenue).number_format = (
                "$#,##0.00"
            )
            ws_detail.cell(row=row, column=4, value=state.threshold_percentage / 100).number_format = (
                "0.0%"
            )
            ws_detail.cell(row=row, column=5, value=state.total_transactions)
            ws_detail.cell(row=row, column=6, value=state.breach_type or "")
            ws_detail.cell(
                row=row,
                column=7,
                value=state.breach_date.strftime("%Y-%m-%d") if state.breach_date else "",
            )
            ws_detail.cell(row=row, column=8, value=state.tax_rate / 100).number_format = "0.00%"
            ws_detail.cell(row=row, column=9, value=state.estimated_liability).number_format = (
                "$#,##0.00"
            )
            row += 1

        # Auto-size columns
        for ws in [ws_summary, ws_detail]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save
        wb.save(output_path)

    def _filter_transactions(
        self,
        transactions: List[Transaction],
        year: int,
        ignore_marketplace: bool,
        include_negatives: bool,
        warnings: List[str],
    ) -> List[Transaction]:
        """Filter transactions based on options."""
        filtered = []

        for txn in transactions:
            # Filter by year
            if txn.date.year != year:
                continue

            # Filter marketplace
            if ignore_marketplace and txn.revenue_type == "marketplace":
                continue

            # Filter negatives
            if not include_negatives and txn.amount < 0:
                continue

            # Validate state code
            if not self._is_valid_state_code(txn.state_code):
                warnings.append(f"Invalid state code: {txn.state_code}")
                continue

            filtered.append(txn)

        if ignore_marketplace:
            filtered_count = len(transactions) - len(filtered)
            if filtered_count > 0:
                warnings.append(f"Filtered out {filtered_count} marketplace transactions")

        return filtered

    def _group_by_state(
        self, transactions: List[Transaction]
    ) -> dict[str, List[Transaction]]:
        """Group transactions by state."""
        by_state: dict[str, List[Transaction]] = {}

        for txn in transactions:
            if txn.state_code not in by_state:
                by_state[txn.state_code] = []
            by_state[txn.state_code].append(txn)

        return by_state

    def _analyze_state(
        self, state_code: str, transactions: List[Transaction], rule: StateRule
    ) -> StateResult:
        """Analyze transactions for a single state."""
        # Sort by date
        sorted_txns = sorted(transactions, key=lambda t: (t.date, t.id))

        # Calculate cumulative breach
        total_revenue = 0.0
        total_txns = 0
        breach_idx = -1
        breach_type = None

        # No threshold = no breach possible (states without sales tax)
        if rule.amount == 0:
            total_revenue = sum(t.amount for t in sorted_txns)
            total_txns = len(sorted_txns)
        else:
            for idx, txn in enumerate(sorted_txns):
                total_revenue += txn.amount
                total_txns += 1

                # Check revenue threshold
                if total_revenue >= rule.amount:
                    breach_idx = idx
                    breach_type = "revenue"
                    break

                # Check transaction threshold
                if rule.transactions and total_txns >= rule.transactions:
                    breach_idx = idx
                    breach_type = "transactions"
                    break

        # Calculate threshold percentage
        if rule.amount > 0:
            revenue_pct = (total_revenue / rule.amount) * 100
            txn_pct = 0.0
            if rule.transactions and rule.transactions > 0:
                txn_pct = (total_txns / rule.transactions) * 100
            threshold_pct = max(revenue_pct, txn_pct)
        else:
            threshold_pct = 0.0

        # Get tax rate
        tax_rate = self.loader.get_tax_rate(state_code) or 0.0

        # Calculate liability (simple estimate: total_revenue * tax_rate)
        liability = 0.0
        if breach_idx >= 0:
            # Only liable for post-nexus revenue
            post_nexus_revenue = sum(t.amount for t in sorted_txns[breach_idx:])
            liability = post_nexus_revenue * (tax_rate / 100)

        return StateResult(
            state_code=state_code,
            has_nexus=breach_idx >= 0,
            total_revenue=total_revenue,
            total_transactions=total_txns,
            threshold_revenue=rule.amount,
            threshold_transactions=rule.transactions,
            breach_type=breach_type,
            breach_date=sorted_txns[breach_idx].date if breach_idx >= 0 else None,
            breach_transaction_id=sorted_txns[breach_idx].id if breach_idx >= 0 else None,
            threshold_percentage=threshold_pct,
            tax_rate=tax_rate,
            estimated_liability=liability,
        )

    def _is_valid_state_code(self, code: str) -> bool:
        """Check if a state code is valid."""
        return len(code) == 2 and code.isalpha()

    def _map_state_name(self, name: str) -> str:
        """Map full state name to 2-letter code."""
        mapping = {
            "ALABAMA": "AL",
            "ALASKA": "AK",
            "ARIZONA": "AZ",
            "ARKANSAS": "AR",
            "CALIFORNIA": "CA",
            "COLORADO": "CO",
            "CONNECTICUT": "CT",
            "DELAWARE": "DE",
            "FLORIDA": "FL",
            "GEORGIA": "GA",
            "HAWAII": "HI",
            "IDAHO": "ID",
            "ILLINOIS": "IL",
            "INDIANA": "IN",
            "IOWA": "IA",
            "KANSAS": "KS",
            "KENTUCKY": "KY",
            "LOUISIANA": "LA",
            "MAINE": "ME",
            "MARYLAND": "MD",
            "MASSACHUSETTS": "MA",
            "MICHIGAN": "MI",
            "MINNESOTA": "MN",
            "MISSISSIPPI": "MS",
            "MISSOURI": "MO",
            "MONTANA": "MT",
            "NEBRASKA": "NE",
            "NEVADA": "NV",
            "NEW HAMPSHIRE": "NH",
            "NEW JERSEY": "NJ",
            "NEW MEXICO": "NM",
            "NEW YORK": "NY",
            "NORTH CAROLINA": "NC",
            "NORTH DAKOTA": "ND",
            "OHIO": "OH",
            "OKLAHOMA": "OK",
            "OREGON": "OR",
            "PENNSYLVANIA": "PA",
            "RHODE ISLAND": "RI",
            "SOUTH CAROLINA": "SC",
            "SOUTH DAKOTA": "SD",
            "TENNESSEE": "TN",
            "TEXAS": "TX",
            "UTAH": "UT",
            "VERMONT": "VT",
            "VIRGINIA": "VA",
            "WASHINGTON": "WA",
            "WEST VIRGINIA": "WV",
            "WISCONSIN": "WI",
            "WYOMING": "WY",
            "DISTRICT OF COLUMBIA": "DC",
        }
        return mapping.get(name.upper(), name[:2])
